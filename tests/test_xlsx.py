import os
from openpyxl import load_workbook

# TODO оформить в тест, добавить ассерты и использовать универсальный путь

XLSX_PATH = os.path.join(os.path.abspath('resources'), 'file_example_XLSX_50.xlsx')


def test_open_xlsx():
    workbook = load_workbook(XLSX_PATH)
    sheet = workbook.active

    assert sheet.cell(row=3, column=2).value == 'Mara'
    assert sheet.max_row == 51
    assert sheet.max_column == 8
    assert workbook.sheetnames[0] == 'Sheet1'
